"""
患者就诊数据汇总导出服务
从 TEMP_PAT_VISIT_LIST 临时表出发，关联业务库各表和应用库 PushLog，
导出每个患者的全部记录到动态列 Excel。
"""
import io
import logging
from datetime import datetime
from typing import Any

from sqlalchemy.orm import Session

from app.models import PushLog

logger = logging.getLogger(__name__)

# 每个类别的列前缀和字段定义
_CATEGORY_DEFS = {
    "surgery": {
        "label": "手术",
        "fields": [
            ("手术名称", "operation_name"),
            ("手术日期", "operation_date"),
            ("麻醉方式", "anesthesia_method"),
            ("手术级别", "operation_level"),
            ("切口愈合等级", "wound_healing_grade"),
        ],
    },
    "progress": {
        "label": "病程",
        "fields": [
            ("记录时间", "event_time"),
            ("标题", "record_name"),
            ("内容", "content"),
            ("创建人", "creator"),
        ],
    },
    "nursing": {
        "label": "护理",
        "fields": [
            ("记录时间", "event_time"),
            ("类型", "record_name"),
            ("内容", "content"),
            ("记录人", "recorder"),
        ],
    },
    "lab": {
        "label": "检验",
        "fields": [
            ("报告时间", "result_time"),
            ("检验项目", "test_name"),
            ("结果明细", "result"),
        ],
    },
    "exam": {
        "label": "检查",
        "fields": [
            ("报告时间", "report_time"),
            ("检查类别", "exam_class"),
            ("描述", "description"),
            ("印象", "impression"),
        ],
    },
    "discharge": {
        "label": "出院记录",
        "fields": [
            ("记录时间", "event_time"),
            ("病历名称", "record_name"),
            ("内容", "content"),
            ("创建科室", "dept"),
            ("创建人", "creator"),
        ],
    },
}

# 基础信息列（每个患者固定）
_BASE_COLUMNS = [
    ("患者ID", 15),
    ("住院号", 12),
    ("住院次数", 8),
    ("患者姓名", 10),
    ("性别", 6),
    ("年龄", 6),
    ("入院日期", 12),
    ("出院日期", 12),
    ("入院科室", 15),
    ("出院科室", 15),
    ("入院诊断", 30),
    ("出院主诊断", 30),
]


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "read"):
        try:
            return str(value.read()).strip()
        except Exception:
            return ""
    return str(value).strip()


_EXCEL_CELL_MAX_LEN = 32767


def _excel_cell_value(value: Any, *, is_datetime_field: bool = False) -> str:
    """生成 Excel 单元格值，自动截断到 32767 字符以内。"""
    text = _format_dt(value) if is_datetime_field else _safe_text(value)
    if len(text) > _EXCEL_CELL_MAX_LEN:
        return text[:_EXCEL_CELL_MAX_LEN]
    return text


def _format_category_record(defn: dict, record: dict) -> str:
    """将一个动态类别记录合并为单个 Excel 单元格文本。"""
    if not record:
        return ""
    lines = []
    for field_label, field_key in defn.get("fields", []):
        raw_value = record.get(field_key, "")
        is_datetime_field = "time" in field_key or "date" in field_key
        value = _excel_cell_value(raw_value, is_datetime_field=is_datetime_field)
        if value:
            lines.append(f"{field_label}：{value}")
    return _excel_cell_value("\n".join(lines))


def _parse_dt(value: Any):
    if isinstance(value, datetime):
        return value
    text = _safe_text(value)
    if not text:
        return None
    for candidate in (text, text[:19], text[:10]):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def _format_dt(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return _safe_text(value)


def _calc_age(birth_date_str, admission_date_str) -> str:
    """从出生日期和入院日期计算年龄。支持 str 和 datetime 类型。"""
    birth = birth_date_str if isinstance(birth_date_str, datetime) else _parse_dt(birth_date_str)
    adm = admission_date_str if isinstance(admission_date_str, datetime) else _parse_dt(admission_date_str)
    if not birth or not adm:
        return ""
    age = adm.year - birth.year - ((adm.month, adm.day) < (birth.month, birth.day))
    return str(age)


def _visit_key(row: dict) -> tuple[str, str]:
    return (
        _safe_text(row.get("patient_id") or row.get("PATIENT_ID") or row.get("患者ID")),
        _safe_text(row.get("visit_number") or row.get("VISIT_NUMBER") or row.get("次数")),
    )


# ---- Oracle 查询 ----

def _execute_oracle_query(conn, sql: str, params: dict = None) -> list[dict]:
    """执行 Oracle 查询，返回字典列表。列名统一用 SQL 中的别名。"""
    cursor = conn.cursor()
    try:
        cursor.execute(sql, params or {})
        columns = [str(col[0]) for col in cursor.description]
        rows = []
        for row in cursor.fetchall():
            item = {}
            for key, value in zip(columns, row):
                if hasattr(value, "read"):
                    value = _safe_text(value)
                item[key] = value
            rows.append(item)
        return rows
    finally:
        cursor.close()


def _query_by_columns(conn, sql: str, params: dict = None) -> list[list]:
    """执行 Oracle 查询，返回 [[col_name, ...], [val, val, ...], ...] 格式。
    用于需要按列位置访问的场景，避免编码问题。"""
    cursor = conn.cursor()
    try:
        cursor.execute(sql, params or {})
        columns = [str(col[0]) for col in cursor.description]
        result = []
        for row in cursor.fetchall():
            result.append(list(row))
        return columns, result
    finally:
        cursor.close()


def _query_patient_list(conn) -> list[dict]:
    """从 TEMP_PAT_VISIT_LIST 查询患者列表。"""
    sql = """
        SELECT t.患者ID, t.住院号, t.住院次数
        FROM TEMP_PAT_VISIT_LIST t
        ORDER BY t.患者ID, t.住院次数
    """
    cols, rows = _query_by_columns(conn, sql)
    result = []
    for row in rows:
        result.append({
            "patient_id": _safe_text(row[0]) if len(row) > 0 else "",
            "admission_no": _safe_text(row[1]) if len(row) > 1 else "",
            "visit_number": _safe_text(row[2]) if len(row) > 2 else "",
        })
    return result


def _query_patient_basic(conn, patient_ids: list[str], visit_numbers: list[str]) -> dict[tuple[str, str], dict]:
    """查询患者基本信息，返回 {(patient_id, visit_number): info_dict}。"""
    if not patient_ids:
        return {}
    pid_binds = ",".join([f":pid{i}" for i in range(len(patient_ids))])
    vn_binds = ",".join([f":vn{i}" for i in range(len(visit_numbers))])
    params = {}
    for i, pid in enumerate(patient_ids):
        params[f"pid{i}"] = pid
    for i, vn in enumerate(visit_numbers):
        params[f"vn{i}"] = vn

    # 用位置映射避免 Oracle 列名编码问题
    sql = f"""
        SELECT 患者ID, 次数, 患者姓名, 性别, 年龄, 出生日期, 入院日期, 出院日期,
               所在科室名称, 入院科室名称, 出院科室名称, 入院诊断, 出院主诊断,
               住院号
        FROM jhemr.v_cybr
        WHERE 患者ID IN ({pid_binds}) AND 次数 IN ({vn_binds})
    """
    try:
        cols, rows = _query_by_columns(conn, sql, params)
    except Exception as exc:
        logger.warning("v_cybr 查询失败: %s", exc)
        return {}

    # 按列位置映射（与 SELECT 顺序一致），避免 Oracle 客户端中文列名编码异常。
    FIELD_MAP = {
        "患者ID": 0, "次数": 1, "患者姓名": 2, "性别": 3, "年龄": 4,
        "出生日期": 5, "入院日期": 6, "出院日期": 7, "所在科室名称": 8,
        "入院科室名称": 9, "出院科室名称": 10, "入院诊断": 11,
        "出院主诊断": 12, "住院号": 13,
    }

    result = {}
    for row in rows:
        pid = _safe_text(row[0]) if len(row) > 0 else ""
        vn = _safe_text(row[1]) if len(row) > 1 else ""
        if not pid:
            continue
        info = {}
        for name, idx in FIELD_MAP.items():
            val = row[idx] if idx < len(row) else ""
            # 保留 datetime 类型，其他转字符串
            if isinstance(val, datetime):
                info[name] = val.strftime("%Y-%m-%d %H:%M:%S")
            else:
                info[name] = _safe_text(val)
        result[(pid, vn)] = info
    return result


def _query_progress_notes(conn, patient_ids: list[str]) -> dict[tuple[str, str], list[dict]]:
    """查询病程记录，按时间+记录人+内容合并。"""
    if not patient_ids:
        return {}
    binds = ",".join([f":p{i}" for i in range(len(patient_ids))])
    params = {f"p{i}": pid for i, pid in enumerate(patient_ids)}
    sql = f"""
        SELECT a.患者ID AS "patient_id", a.次数 AS "visit_number",
               b.病历标题时间 AS "event_time", b.病历名称 AS "record_name",
               b.病历内容 AS "content", b.创建人 AS "creator"
        FROM jhemr.v_cybr a
        JOIN bcjl202603 b ON a.患者ID = b.患者ID AND a.次数 = b.次数
        WHERE a.患者ID IN ({binds})
        ORDER BY a.患者ID, a.次数, b.病历标题时间
    """
    try:
        rows = _execute_oracle_query(conn, sql, params)
    except Exception as exc:
        logger.warning("病程查询失败: %s", exc)
        return {}
    # 按时间+记录人+内容合并
    result: dict[tuple[str, str], list[dict]] = {}
    for r in rows:
        key = _visit_key(r)
        time_val = _safe_text(r.get("event_time"))
        creator = _safe_text(r.get("creator"))
        content = _safe_text(r.get("content"))
        merge_key = f"{time_val}|{creator}|{content[:100]}"
        group = result.setdefault(key, [])
        existing = next((g for g in group if g.get("_merge_key") == merge_key), None)
        if existing:
            existing["record_name"] = _safe_text(existing.get("record_name")) + "、" + _safe_text(r.get("record_name"))
        else:
            r["_merge_key"] = merge_key
            group.append(r)
    return result


def _query_nursing_records(conn, patient_ids: list[str]) -> dict[tuple[str, str], list[dict]]:
    """查询护理记录，排除病重病危，按时间+记录人+内容合并。"""
    if not patient_ids:
        return {}
    binds = ",".join([f":p{i}" for i in range(len(patient_ids))])
    params = {f"p{i}": pid for i, pid in enumerate(patient_ids)}
    sql = f"""
        SELECT a.患者ID AS "patient_id", a.次数 AS "visit_number",
               n.护理记录时间 AS "event_time", n.护理单类型 AS "record_name",
               n.病情观察及护理措施 AS "content", n.记录人 AS "recorder"
        FROM jhemr.v_cybr a
        JOIN ydhl_202501 n ON n.患者ID = a.患者ID || '_' || a.次数
        WHERE a.患者ID IN ({binds})
          AND NVL(n.护理单类型, ' ') NOT LIKE '%病重%'
          AND NVL(n.护理单类型, ' ') NOT LIKE '%病危%'
        ORDER BY a.患者ID, a.次数, n.护理记录时间
    """
    try:
        rows = _execute_oracle_query(conn, sql, params)
    except Exception as exc:
        logger.warning("护理查询失败: %s", exc)
        return {}
    # 按时间+记录人+内容合并
    result: dict[tuple[str, str], list[dict]] = {}
    for r in rows:
        key = _visit_key(r)
        time_val = _safe_text(r.get("event_time"))
        recorder = _safe_text(r.get("recorder"))
        content = _safe_text(r.get("content"))
        merge_key = f"{time_val}|{recorder}|{content[:100]}"
        group = result.setdefault(key, [])
        existing = next((g for g in group if g.get("_merge_key") == merge_key), None)
        if existing:
            existing["record_name"] = _safe_text(existing.get("record_name")) + "、" + _safe_text(r.get("record_name"))
        else:
            r["_merge_key"] = merge_key
            group.append(r)
    return result


def _query_lab_reports(conn, patient_ids: list[str]) -> dict[tuple[str, str], list[dict]]:
    """查询检验报告，按 TEST_NO 合并同一检验单的所有子项。"""
    if not patient_ids:
        return {}
    binds = ",".join([f":p{i}" for i in range(len(patient_ids))])
    params = {f"p{i}": pid for i, pid in enumerate(patient_ids)}
    sql = f"""
        SELECT a.患者ID AS "patient_id", a.次数 AS "visit_number",
               lm.TEST_NO AS "test_no",
               b.ITEM_NAME AS "item_name",
               c.REPORT_ITEM_NAME AS "report_item_name", c.RESULT AS "result",
               c.UNITS AS "units",
               CASE c.ABNORMAL_INDICATOR WHEN 'L' THEN '低' WHEN 'H' THEN '高' ELSE c.ABNORMAL_INDICATOR END AS "abnormal_indicator",
               c.PRINT_CONTEXT AS "reference_range",
               c.RESULT_DATE_TIME AS "result_time"
        FROM jhemr.v_cybr a
        INNER JOIN his.LAB_TEST_MASTER lm ON a.患者ID = lm.PATIENT_ID AND a.次数 = lm.VISIT_ID
        INNER JOIN his.LAB_TEST_ITEMS b ON lm.TEST_NO = b.TEST_NO
        INNER JOIN his.LAB_RESULT c ON c.TEST_NO = lm.TEST_NO AND c.ITEM_NO = b.ITEM_NO
        WHERE a.患者ID IN ({binds})
        ORDER BY a.患者ID, a.次数, lm.TEST_NO, c.RESULT_DATE_TIME
    """
    try:
        rows = _execute_oracle_query(conn, sql, params)
    except Exception as exc:
        logger.warning("检验查询失败: %s", exc)
        return {}
    # 按 TEST_NO 合并
    result: dict[tuple[str, str], list[dict]] = {}
    groups: dict[tuple[str, str], dict[str, list]] = {}
    for r in rows:
        vkey = _visit_key(r)
        test_no = _safe_text(r.get("test_no"))
        groups.setdefault(vkey, {}).setdefault(test_no, []).append(r)

    for vkey, test_groups in groups.items():
        merged = []
        for test_no, items in test_groups.items():
            if not items:
                continue
            first = items[0]
            item_lines = []
            for it in items:
                name = _safe_text(it.get("report_item_name") or it.get("item_name"))
                val = _safe_text(it.get("result"))
                unit = _safe_text(it.get("units"))
                abn = _safe_text(it.get("abnormal_indicator"))
                ref = _safe_text(it.get("reference_range"))
                line = f"{name}: {val}"
                if unit:
                    line += f" {unit}"
                if abn:
                    line += f" ({abn})"
                if ref:
                    line += f" [参考:{ref}]"
                item_lines.append(line)
            merged.append({
                "test_no": test_no,
                "test_name": _safe_text(first.get("item_name")),
                "result": "\n".join(item_lines),
                "result_time": _safe_text(first.get("result_time")),
            })
        result[vkey] = merged
    return result


def _query_exam_reports(conn, patient_ids: list[str]) -> dict[tuple[str, str], list[dict]]:
    """查询检查报告，按 EXAM_NO 合并。"""
    if not patient_ids:
        return {}
    binds = ",".join([f":p{i}" for i in range(len(patient_ids))])
    params = {f"p{i}": pid for i, pid in enumerate(patient_ids)}
    sql = f"""
        SELECT a.患者ID AS "patient_id", a.次数 AS "visit_number",
               em.EXAM_NO AS "exam_no", em.EXAM_CLASS AS "exam_class",
               er.DESCRIPTION AS "description", er.IMPRESSION AS "impression",
               NVL(er.REPORT_TIME, em.EXAM_DATE_TIME) AS "report_time"
        FROM jhemr.v_cybr a
        JOIN his.EXAM_MASTER em ON a.患者ID = em.PATIENT_ID AND a.次数 = em.VISIT_ID
        LEFT JOIN his.EXAM_REPORT er ON em.EXAM_NO = er.EXAM_NO
        WHERE a.患者ID IN ({binds})
        ORDER BY a.患者ID, a.次数, em.EXAM_NO
    """
    try:
        rows = _execute_oracle_query(conn, sql, params)
    except Exception as exc:
        logger.warning("检查查询失败: %s", exc)
        return {}
    # 按 EXAM_NO 合并
    result: dict[tuple[str, str], list[dict]] = {}
    groups: dict[tuple[str, str], dict[str, dict]] = {}
    for r in rows:
        vkey = _visit_key(r)
        exam_no = _safe_text(r.get("exam_no"))
        g = groups.setdefault(vkey, {})
        if exam_no not in g:
            g[exam_no] = {
                "exam_no": exam_no,
                "exam_class": _safe_text(r.get("exam_class")),
                "description": _safe_text(r.get("description")),
                "impression": _safe_text(r.get("impression")),
                "report_time": _safe_text(r.get("report_time")),
            }
        else:
            existing = g[exam_no]
            desc = _safe_text(r.get("description"))
            imp = _safe_text(r.get("impression"))
            if desc and desc not in existing["description"]:
                existing["description"] += "\n" + desc
            if imp and imp not in existing["impression"]:
                existing["impression"] += "\n" + imp

    for vkey, exam_groups in groups.items():
        result[vkey] = list(exam_groups.values())
    return result


def _query_discharge_records(conn, patient_ids: list[str]) -> dict[tuple[str, str], list[dict]]:
    """查询出院记录，返回 {(patient_id, visit_number): [records]}。
    
    数据源：jhemr.V_cyJL（病历文书表）
    过滤条件：病历名称 LIKE '%出院记录%'
    """
    if not patient_ids:
        return {}
    binds = ",".join([f":p{i}" for i in range(len(patient_ids))])
    params = {f"p{i}": pid for i, pid in enumerate(patient_ids)}
    sql = f"""
        SELECT a.患者ID AS "patient_id", a.次数 AS "visit_number",
               a.病历创建时间 AS "event_time",
               a.病历名称 AS "record_name",
               a.病历内容 AS "content",
               a.病历创建所在科室 AS "dept",
               a.创建人 AS "creator"
        FROM jhemr.V_cyJL a
        WHERE a.患者ID IN ({binds})
          AND a.病历名称 LIKE '%出院记录%'
          AND a.RN = 1
    """
    try:
        cols, rows = _query_by_columns(conn, sql, params)
    except Exception as exc:
        logger.warning("出院记录查询失败: %s", exc)
        return {}

    # 按位置映射：0=patient_id, 1=visit_number, 2=event_time, 3=record_name, 4=content, 5=dept, 6=creator
    result: dict[tuple[str, str], list[dict]] = {}
    for row in rows:
        pid = _safe_text(row[0]) if len(row) > 0 else ""
        vn = _safe_text(row[1]) if len(row) > 1 else ""
        if not pid:
            continue
        key = (pid, vn)
        result.setdefault(key, []).append({
            "event_time": _format_dt(row[2]) if len(row) > 2 else "",
            "record_name": _safe_text(row[3]) if len(row) > 3 else "",
            "content": _safe_text(row[4]) if len(row) > 4 else "",
            "dept": _safe_text(row[5]) if len(row) > 5 else "",
            "creator": _safe_text(row[6]) if len(row) > 6 else "",
        })
    return result


def _query_discharge_records_from_emr(emr_cfg: dict, patient_keys: list[tuple[str, str]]) -> dict[tuple[str, str], list[dict]]:
    """从海量库查询出院记录。异常时由调用方决定是否回退 Oracle。"""
    from app.emr_vastbase_client import fetch_emr_documents_by_visits
    raw = fetch_emr_documents_by_visits(emr_cfg, patient_keys, document_kind="discharge")
    result: dict[tuple[str, str], list[dict]] = {}
    for key, records in raw.items():
        result[key] = [
            {
                "event_time": r.get("event_time", ""),
                "record_name": r.get("record_name", ""),
                "content": r.get("content", ""),
                "dept": r.get("dept", ""),
                "creator": r.get("creator", ""),
            }
            for r in records
        ]
    return result


def _query_progress_notes_from_emr(emr_cfg: dict, patient_keys: list[tuple[str, str]]) -> dict[tuple[str, str], list[dict]]:
    """从海量库查询病程记录（排除出院记录），按时间+创建人+内容合并。异常时由调用方决定是否回退 Oracle。"""
    from app.emr_vastbase_client import fetch_emr_documents_by_visits
    raw = fetch_emr_documents_by_visits(emr_cfg, patient_keys, document_kind="progress")
    result: dict[tuple[str, str], list[dict]] = {}
    for key, records in raw.items():
        merged: list[dict] = []
        for r in records:
            time_val = r.get("event_time", "")
            creator = r.get("creator", "")
            content = r.get("content", "")
            merge_key = f"{time_val}|{creator}|{content[:100]}"
            existing = next((g for g in merged if g.get("_merge_key") == merge_key), None)
            if existing:
                existing["record_name"] = existing.get("record_name", "") + "、" + r.get("record_name", "")
            else:
                r["_merge_key"] = merge_key
                merged.append(r)
        for m in merged:
            m.pop("_merge_key", None)
        result[key] = [
            {
                "event_time": r.get("event_time", ""),
                "record_name": r.get("record_name", ""),
                "content": r.get("content", ""),
                "creator": r.get("creator", ""),
            }
            for r in merged
        ]
    return result


def _query_frontpage_surgery(conn, patient_ids: list[str]) -> dict[tuple[str, str], list[dict]]:
    """查询病案首页手术，返回 {(patient_id, visit_number): [records]}。"""
    if not patient_ids:
        return {}
    binds = ",".join([f":p{i}" for i in range(len(patient_ids))])
    params = {f"p{i}": pid for i, pid in enumerate(patient_ids)}
    sql = f"""
        SELECT 患者ID, 次数,
               手术1, 手术日期1, 手术2, 手术日期2, 手术3, 手术日期3,
               手术4, 手术日期4, 手术5, 手术日期5
        FROM jhemr.v_cybr
        WHERE 患者ID IN ({binds})
    """
    try:
        cols, rows = _query_by_columns(conn, sql, params)
    except Exception as exc:
        logger.warning("首页手术查询失败: %s", exc)
        return {}

    # 按位置映射：0=患者ID, 1=次数, 2=手术1, 3=手术日期1, ..., 11=手术日期5
    result: dict[tuple[str, str], list[dict]] = {}
    for row in rows:
        pid = _safe_text(row[0]) if len(row) > 0 else ""
        vn = _safe_text(row[1]) if len(row) > 1 else ""
        if not pid:
            continue
        surgeries = []
        for i in range(5):
            raw = _safe_text(row[2 + i * 2]) if (2 + i * 2) < len(row) else ""
            date = _safe_text(row[3 + i * 2]) if (3 + i * 2) < len(row) else ""
            if raw:
                surgeries.append({"operation_name": raw, "operation_date": date, "raw_text": raw})
        if surgeries:
            result.setdefault((pid, vn), []).extend(surgeries)
    return result


# ---- PushLog 查询（应用库）----

def _query_push_logs(db: Session, patient_keys: list[tuple[str, str]]) -> dict[tuple[str, str], list[dict]]:
    """查询 PushLog 的 request_json 和 response_json。"""
    if not patient_keys:
        return {}
    from sqlalchemy import or_
    conditions = []
    for pid, vn in patient_keys:
        conditions.append(
            (PushLog.patient_id == pid) & (PushLog.visit_number == vn)
        )
    logs = db.query(PushLog).filter(
        PushLog.status == "success",
        or_(*conditions),
    ).order_by(PushLog.push_time.desc()).all()

    result: dict[tuple[str, str], list[dict]] = {}
    for log in logs:
        key = (_safe_text(log.patient_id), _safe_text(log.visit_number))
        result.setdefault(key, []).append({
            "push_time": _format_dt(log.push_time),
            "audit_type_code": _safe_text(log.audit_type_code),
            "request_json": _safe_text(log.request_json or ""),
            "response_json": _safe_text(log.response_json or ""),
        })
    return result


# ---- Excel 构建 ----

def _build_excel(patient_data: list[dict]) -> bytes:
    """构建动态列 Excel。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl 未安装，无法导出 Excel")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "患者就诊数据汇总"

    # 计算每个类别的最大记录数
    max_counts = {}
    for cat in _CATEGORY_DEFS:
        max_counts[cat] = max((len(p.get(cat, [])) for p in patient_data), default=0)
        if max_counts[cat] == 0:
            max_counts[cat] = 0

    # 构建表头
    headers = []
    col_widths = []
    for name, width in _BASE_COLUMNS:
        headers.append(name)
        col_widths.append(width)
    for cat, defn in _CATEGORY_DEFS.items():
        count = max_counts.get(cat, 0)
        if count == 0:
            continue
        label = defn["label"]
        for i in range(1, count + 1):
            col_name = f"{label}{i}"
            headers.append(col_name)
            col_widths.append(45)

    # 写表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 写数据行
    data_font = Font(size=9)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row_idx, p in enumerate(patient_data, 2):
        col = 1
        # 基础列
        for name, _ in _BASE_COLUMNS:
            cell = ws.cell(row=row_idx, column=col, value=_excel_cell_value(p.get(name, "")))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            col += 1
        # 动态列
        for cat, defn in _CATEGORY_DEFS.items():
            records = p.get(cat, [])
            count = max_counts.get(cat, 0)
            for i in range(count):
                rec = records[i] if i < len(records) else {}
                val = _format_category_record(defn, rec)
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                col += 1

    # 冻结首行
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- 主入口 ----

def export_patient_visit_summary(db: Session) -> tuple[bytes, str]:
    """导出患者就诊数据汇总 Excel。"""
    from app.config import load_config
    from app.oracle_client import get_oracle_connection
    from app.services.config_parser import ConfigParser

    config = load_config()
    oracle_cfg = ConfigParser.parse_oracle_config(config)
    emr_cfg = ConfigParser.parse_emr_vastbase_config(config)
    emr_enabled = bool(emr_cfg.get("enabled"))
    conn = get_oracle_connection(oracle_cfg)

    try:
        # 1. 查询临时表患者列表
        patients = _query_patient_list(conn)
        if not patients:
            raise ValueError("TEMP_PAT_VISIT_LIST 中没有数据")

        patient_ids = list({_safe_text(p.get("patient_id")) for p in patients})
        visit_numbers = list({_safe_text(p.get("visit_number")) for p in patients})
        patient_keys = [
            (_safe_text(p.get("patient_id")), _safe_text(p.get("visit_number")))
            for p in patients
        ]

        # 2. 查询各表
        basic_info = _query_patient_basic(conn, patient_ids, visit_numbers)
        nursing = _query_nursing_records(conn, patient_ids)
        lab = _query_lab_reports(conn, patient_ids)
        exam = _query_exam_reports(conn, patient_ids)
        surgery = _query_frontpage_surgery(conn, patient_ids)

        # 2a. 病程记录：海量库优先，异常时回退 Oracle
        progress: dict[tuple[str, str], list[dict]] = {}
        if emr_enabled and emr_cfg.get("use_for_export_progress", True):
            try:
                progress = _query_progress_notes_from_emr(emr_cfg, patient_keys)
                logger.info("病程记录来源: 海量库, %d 患者住院次", len(progress))
            except Exception as exc:
                if emr_cfg.get("fallback_to_oracle", True):
                    logger.warning("海量库病程查询失败，回退 Oracle: %s", exc)
                    progress = _query_progress_notes(conn, patient_ids)
                    logger.info("病程记录来源: Oracle 回退, %d 患者住院次", len(progress))
                else:
                    logger.error("海量库病程查询失败且未启用回退: %s", exc)
                    raise
        else:
            progress = _query_progress_notes(conn, patient_ids)

        # 2b. 出院记录：海量库优先，异常时回退 Oracle
        discharge: dict[tuple[str, str], list[dict]] = {}
        if emr_enabled and emr_cfg.get("use_for_export_discharge", True):
            try:
                discharge = _query_discharge_records_from_emr(emr_cfg, patient_keys)
                logger.info("出院记录来源: 海量库, %d 患者住院次", len(discharge))
            except Exception as exc:
                if emr_cfg.get("fallback_to_oracle", True):
                    logger.warning("海量库出院记录查询失败，回退 Oracle: %s", exc)
                    discharge = _query_discharge_records(conn, patient_ids)
                    logger.info("出院记录来源: Oracle 回退, %d 患者住院次", len(discharge))
                else:
                    logger.error("海量库出院记录查询失败且未启用回退: %s", exc)
                    raise
        else:
            discharge = _query_discharge_records(conn, patient_ids)

        # 3. 查询 PushLog
        push_logs = _query_push_logs(db, patient_keys)

        # 4. 组装每行数据
        patient_data = []
        for p in patients:
            pid = _safe_text(p.get("patient_id"))
            vn = _safe_text(p.get("visit_number"))
            adm_no = _safe_text(p.get("admission_no"))
            key = (pid, vn)
            info = basic_info.get(key, {})

            birth_date = info.get("出生日期") or ""
            admission_date = info.get("入院日期") or ""
            age = info.get("年龄") or _calc_age(birth_date, admission_date)

            # PushLog：取最新一条
            logs = push_logs.get(key, [])
            latest_log = logs[0] if logs else {}

            row = {
                "患者ID": pid,
                "住院号": adm_no,
                "住院次数": vn,
                "患者姓名": info.get("患者姓名") or "",
                "性别": info.get("性别") or "",
                "年龄": age,
                "入院日期": info.get("入院日期") or "",
                "出院日期": info.get("出院日期") or "",
                "入院科室": info.get("入院科室名称") or info.get("所在科室名称") or "",
                "出院科室": info.get("出院科室名称") or "",
                "入院诊断": info.get("入院诊断") or "",
                "出院主诊断": info.get("出院主诊断") or "",
                "surgery": surgery.get(key, []),
                "progress": progress.get(key, []),
                "nursing": nursing.get(key, []),
                "lab": lab.get(key, []),
                "exam": exam.get(key, []),
                "discharge": discharge.get(key, []),
            }

            # 添加 PushLog 的 request_json / response_json（多条）
            if logs:
                for i, lg in enumerate(logs[:5]):  # 最多5条
                    row[f"推送时间{i+1}"] = lg.get("push_time", "")
                    row[f"审计类型{i+1}"] = lg.get("audit_type_code", "")
                    row[f"推送JSON{i+1}"] = lg.get("request_json", "")[:32000]  # Excel 单元格限制
                    row[f"返回JSON{i+1}"] = lg.get("response_json", "")[:32000]

            patient_data.append(row)

        # 5. 构建 Excel（PushLog 列也加入动态列）
        # 先计算 PushLog 最大列数
        max_push = max((len(push_logs.get((_safe_text(p.get("patient_id")), _safe_text(p.get("visit_number"))), [])) for p in patients), default=0)
        max_push = min(max_push, 5)

        # 重新构建 Excel（包含 PushLog 动态列）
        xlsx_bytes = _build_excel_with_pushlog(patient_data, max_push)
        return xlsx_bytes, "xlsx"

    finally:
        try:
            conn.close()
        except Exception:
            pass


def _build_excel_with_pushlog(patient_data: list[dict], max_push: int) -> bytes:
    """构建包含 PushLog 动态列的 Excel。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl 未安装")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "患者就诊数据汇总"

    # 计算每个类别的最大记录数
    max_counts = {}
    for cat in _CATEGORY_DEFS:
        max_counts[cat] = max((len(p.get(cat, [])) for p in patient_data), default=0)

    # 构建表头
    headers = []
    col_widths = []
    for name, width in _BASE_COLUMNS:
        headers.append(name)
        col_widths.append(width)
    for cat, defn in _CATEGORY_DEFS.items():
        count = max_counts.get(cat, 0)
        for i in range(count):
            headers.append(f"{defn['label']}{i+1}")
            col_widths.append(45)
    # PushLog 动态列
    pushlog_fields = ["推送时间", "审计类型", "推送JSON", "返回JSON"]
    for i in range(max_push):
        for field in pushlog_fields:
            headers.append(f"质控{i+1}_{field}")
            col_widths.append(40 if "JSON" in field else 16)

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 写数据
    data_font = Font(size=9)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row_idx, p in enumerate(patient_data, 2):
        col = 1
        # 基础列
        for name, _ in _BASE_COLUMNS:
            cell = ws.cell(row=row_idx, column=col, value=_excel_cell_value(p.get(name, "")))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            col += 1
        # 动态列
        for cat, defn in _CATEGORY_DEFS.items():
            records = p.get(cat, [])
            count = max_counts.get(cat, 0)
            for i in range(count):
                rec = records[i] if i < len(records) else {}
                val = _format_category_record(defn, rec)
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                col += 1
        # PushLog 动态列
        for i in range(max_push):
            for field in pushlog_fields:
                key = f"{field}{i+1}"
                val = p.get(key, "")
                cell = ws.cell(row=row_idx, column=col, value=_excel_cell_value(val))
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                col += 1

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
