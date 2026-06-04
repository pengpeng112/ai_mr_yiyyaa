"""
质控反馈导出服务
支持 CSV/Excel 导出病例详情视图
"""
import csv
import io
import json
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Iterable, Optional

from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session

from app.config import load_config
from app.models import (
    AuditConclusion,
    AuditDimensionResult,
    Department,
    PushLog,
    QCFeedback,
    QCFeedbackHistory,
    User,
)
from app.services.patient_snapshot import (
    apply_privacy_masking,
    extract_patient_snapshot,
    extract_raw_record_sections,
    normalize_privacy_masking_config,
)
from app.services.audit_type_registry import AuditTypeRegistry


class _ExportBytes(bytes):
    """导出内容，兼容 bytes 用法和 (bytes, format) 解包。"""

    def __new__(cls, data: bytes, file_format: str):
        obj = super().__new__(cls, data)
        obj.file_format = file_format
        return obj

    def __iter__(self):
        yield bytes(self)
        yield self.file_format


class FeedbackExportService:
    """反馈导出服务（按病例详情维度导出）"""

    # 0 表示不限制导出行数（导出全部）
    EXPORT_MAX_ROWS = 0
    CANONICAL_DIMENSION_COLUMNS = [
        ("诊断一致性", ["诊断一致性"]),
        ("护理级别执行一致性", ["优先护理级别执行一致性", "护理级别一致性"]),
        ("生命体征一致性", ["生命体征一致性"]),
        ("病情描述一致性", ["病情描述一致性"]),
        ("治疗措施执行一致性", ["优先治疗措施执行一致性", "诊疗措施一致性"]),
        ("时间线一致性", ["优先时间线一致性", "时间记录一致性"]),
    ]
    LAB_EXAM_AUDIT_CODES = {
        "jyjc_vs_bcnursing",
        "lab_exam_vs_progress_nursing",
        "lab_exam_structured_progress_nursing",
        "labexam_vs_progress",
    }
    LAB_EXAM_EXCLUDED_DIMENSION_COLUMNS = {
        "诊断一致性",
        "护理级别一致",
        "生命体征信息缺失",
        "病情描述一致",
        "诊疗措施一致",
        "时间记录合理",
        "护理级别一致性",
        "生命体征一致性",
        "病情一致性",
        "治疗措施一致性",
        "时间线一致性",
    }
    ORACLE_IN_CHUNK_SIZE = 900

    def __init__(self, db: Session):
        self.db = db
        self.last_export_count = 0

    def _iter_chunks(self, values: Iterable, chunk_size: int | None = None):
        size = chunk_size or self.ORACLE_IN_CHUNK_SIZE
        if size <= 0:
            size = self.ORACLE_IN_CHUNK_SIZE
        items = list(values or [])
        for idx in range(0, len(items), size):
            yield items[idx: idx + size]

    def _parse_focus_items(self, raw_text: str) -> list[str]:
        if not raw_text:
            return []
        try:
            value = json.loads(raw_text)
            if isinstance(value, list):
                return [str(item).strip() for item in value if str(item).strip()]
        except Exception:
            return []
        return []

    def _status_label(self, status_value: str) -> str:
        labels = {
            "pending": "待处理",
            "acknowledged": "已确认",
            "rectified": "已整改",
            "closed": "已关闭",
        }
        return labels.get(status_value or "", status_value or "")

    def _severity_label(self, severity_value: str) -> str:
        labels = {
            "high": "高",
            "medium": "中",
            "low": "低",
        }
        return labels.get(severity_value or "", severity_value or "")

    def _alert_level_label(self, alert_level: str) -> str:
        labels = {
            "red": "红灯",
            "yellow": "黄灯",
            "blue": "蓝灯",
            "gray": "灰灯",
        }
        return labels.get(alert_level or "", alert_level or "")

    def _push_strategy_label(self, strategy: str) -> str:
        labels = {
            "immediate": "立即推送",
            "batch": "批量汇总",
            "shift_summary": "交班汇总",
            "review_only": "仅复核",
        }
        return labels.get(strategy or "", strategy or "")

    def _outcome_bucket_label(self, bucket: str) -> str:
        labels = {
            "primary": "主要问题",
            "secondary": "次要问题",
            "none": "无问题",
        }
        return labels.get(bucket or "", bucket or "")

    def _normalize_dimension_name(self, value: str) -> str:
        name = str(value or "").strip()
        if not name:
            return "未命名维度"
        return name

    def _merge_dimension_value(self, values: list[str], sep: str) -> str:
        cleaned = [str(v or "").strip() for v in values if str(v or "").strip()]
        if not cleaned:
            return ""
        return sep.join(cleaned)

    def _pick_canonical_dimension_value(self, item_dims: dict, prefixes: list[str]) -> dict:
        if not item_dims:
            return {"status": "", "medical": "", "nursing": "", "explanation": ""}

        matched_values = []
        for prefix in prefixes:
            prefix_text = str(prefix or "").strip()
            if not prefix_text:
                continue
            current_matches = []
            for dim_name, dim_value in item_dims.items():
                if str(dim_name or "").strip().startswith(prefix_text):
                    current_matches.append(dim_value or {})
            if current_matches:
                matched_values = current_matches
                break

        if not matched_values:
            return {"status": "", "medical": "", "nursing": "", "explanation": ""}

        return {
            "status": self._merge_dimension_value([v.get("status", "") for v in matched_values], " | "),
            "medical": self._merge_dimension_value([v.get("medical", "") for v in matched_values], "\n---\n"),
            "nursing": self._merge_dimension_value([v.get("nursing", "") for v in matched_values], "\n---\n"),
            "explanation": self._merge_dimension_value([v.get("explanation", "") for v in matched_values], "\n---\n"),
        }

    def _format_canonical_dimension_cell(self, value: dict) -> str:
        dim_value = value or {}
        return "\n".join(
            [
                f"状态：{dim_value.get('status', '')}",
                f"病程记录：{dim_value.get('medical', '')}",
                f"护理记录：{dim_value.get('nursing', '')}",
                f"说明：{dim_value.get('explanation', '')}",
            ]
        )

    def _format_dimension_judgment(self, dim_value: dict) -> str:
        value = dim_value or {}
        lines = [
            f"维度：{value.get('dimension', '')}",
            f"判断：{value.get('status', '')}",
            f"严重度：{value.get('severity', '')}",
            f"预警灯号：{value.get('alert_level', '')}",
            f"闭环时限：{value.get('closure_hours', '')}",
            f"推送策略：{value.get('push_strategy', '')}",
            f"问题分层：{value.get('outcome_bucket', '')}",
            f"说明：{value.get('explanation', '')}",
            f"建议：{value.get('recommendation', '')}",
        ]
        return "\n".join([line for line in lines if not line.endswith("：")])

    def _format_all_dimension_judgments(self, dimensions: list[dict]) -> str:
        items = [self._format_dimension_judgment(item) for item in dimensions or []]
        return "\n---\n".join([item for item in items if item.strip()])

    def _is_lab_exam_export(self, rows: list[dict], audit_type_code: Optional[str]) -> bool:
        audit_code = str(audit_type_code or "").strip()
        if audit_code in self.LAB_EXAM_AUDIT_CODES:
            return True
        for item in rows or []:
            row_code = str(item.get("audit_type_code") or "").strip()
            row_name = str(item.get("audit_type_name") or "").strip()
            if row_code in self.LAB_EXAM_AUDIT_CODES:
                return True
            if ("检验" in row_name or "检查" in row_name) and ("护理" in row_name or "病程" in row_name):
                return True
        return False

    def _should_exclude_dimension_for_export(self, dim_name: str, is_lab_exam_export: bool) -> bool:
        return is_lab_exam_export and self._normalize_dimension_name(dim_name) in self.LAB_EXAM_EXCLUDED_DIMENSION_COLUMNS

    def _dimension_columns_for_export(self, rows: list[dict], audit_type_code: Optional[str]) -> list[tuple[str, list[str]]]:
        audit_code = str(audit_type_code or "").strip()
        if not audit_code or audit_code == "progress_vs_nursing":
            return self.CANONICAL_DIMENSION_COLUMNS

        is_lab_exam_export = self._is_lab_exam_export(rows, audit_type_code)
        dimension_names = []
        seen = set()
        for item in rows or []:
            for dim in item.get("dimension_items") or []:
                dim_name = self._normalize_dimension_name(dim.get("dimension") or dim.get("dimension_code") or "")
                if self._should_exclude_dimension_for_export(dim_name, is_lab_exam_export):
                    continue
                if dim_name and dim_name not in seen:
                    seen.add(dim_name)
                    dimension_names.append(dim_name)
            for dim_name in (item.get("dimensions") or {}).keys():
                normalized_name = self._normalize_dimension_name(dim_name)
                if self._should_exclude_dimension_for_export(normalized_name, is_lab_exam_export):
                    continue
                if normalized_name and normalized_name not in seen:
                    seen.add(normalized_name)
                    dimension_names.append(normalized_name)

        return [(name, [name]) for name in dimension_names]

    def _build_case_rows(
        self,
        role_name: str,
        current_user_dept_id: Optional[int],
        status: Optional[str] = None,
        severity: Optional[str] = None,
        audit_type_code: Optional[str] = None,
        dept_id: Optional[int] = None,
        days: int = 30,
        keyword: Optional[str] = None,
        mask_sensitive: bool = False,
    ) -> list[dict]:
        mask_cfg = normalize_privacy_masking_config(load_config().get("privacy_masking", {})) if mask_sensitive else None
        departments = self.db.query(Department).all()
        dept_by_name = {str(d.name or "").strip(): d for d in departments if str(d.name or "").strip()}
        dept_by_id = {d.id: d for d in departments}

        current_dept_name = None
        if current_user_dept_id:
            current_dept = dept_by_id.get(current_user_dept_id)
            current_dept_name = current_dept.name if current_dept else None

        issue_count_subquery = (
            self.db.query(
                AuditDimensionResult.push_log_id.label("log_id"),
                func.count(AuditDimensionResult.id).label("issue_count"),
            )
            .group_by(AuditDimensionResult.push_log_id)
            .subquery()
        )

        latest_feedback_subquery = (
            self.db.query(
                QCFeedback.push_log_id.label("log_id"),
                func.max(QCFeedback.id).label("feedback_id"),
            )
            .group_by(QCFeedback.push_log_id)
            .subquery()
        )

        query = (
            self.db.query(
                PushLog,
                AuditConclusion,
                QCFeedback,
                issue_count_subquery.c.issue_count,
            )
            .outerjoin(AuditConclusion, AuditConclusion.push_log_id == PushLog.id)
            .outerjoin(latest_feedback_subquery, latest_feedback_subquery.c.log_id == PushLog.id)
            .outerjoin(QCFeedback, QCFeedback.id == latest_feedback_subquery.c.feedback_id)
            .outerjoin(issue_count_subquery, issue_count_subquery.c.log_id == PushLog.id)
            .filter(PushLog.status == "success")
            .filter(PushLog.push_time >= datetime.now() - timedelta(days=days))
        )
        query = query.filter(or_(QCFeedback.id.is_(None), QCFeedback.status != "deleted"))

        if role_name != "admin":
            dept_filters = [QCFeedback.dept_id == current_user_dept_id]
            if current_dept_name:
                dept_filters.append(and_(QCFeedback.id.is_(None), PushLog.dept == current_dept_name))
            query = query.filter(or_(*dept_filters))
        elif dept_id:
            dept_obj = dept_by_id.get(dept_id)
            dept_name = dept_obj.name if dept_obj else None
            query = query.filter(
                or_(
                    QCFeedback.dept_id == dept_id,
                    and_(QCFeedback.id.is_(None), PushLog.dept == dept_name),
                )
            )

        if status:
            if status == "pending":
                query = query.filter(or_(QCFeedback.id.is_(None), QCFeedback.status == "pending"))
            else:
                query = query.filter(QCFeedback.status == status)

        if severity:
            query = query.filter(or_(AuditConclusion.severity == severity, PushLog.severity == severity))

        if audit_type_code:
            audit_code = audit_type_code.strip()
            if audit_code == "progress_vs_nursing":
                query = query.filter(
                    or_(
                        PushLog.audit_type_code == audit_code,
                        PushLog.audit_type_code.is_(None),
                        PushLog.audit_type_code == "",
                    )
                )
            elif audit_code:
                query = query.filter(PushLog.audit_type_code == audit_code)

        if keyword:
            kw = keyword.strip()
            if kw:
                like_pattern = f"%{kw}%"
                query = query.filter(
                    or_(
                        PushLog.patient_id.like(like_pattern),
                        PushLog.patient_name.like(like_pattern),
                        PushLog.admission_no.like(like_pattern),
                    )
                )

        ordered_query = query.order_by(PushLog.push_time.desc())
        if self.EXPORT_MAX_ROWS and self.EXPORT_MAX_ROWS > 0:
            ordered_query = ordered_query.limit(self.EXPORT_MAX_ROWS)
        rows = ordered_query.all()
        if not rows:
            return []

        log_ids = [log.id for log, _, _, _ in rows]
        feedback_ids = [fb.id for _, _, fb, _ in rows if fb]

        dimensions_by_log_id = defaultdict(list)
        if log_ids:
            dims = []
            for id_chunk in self._iter_chunks(log_ids):
                dims.extend(
                    self.db.query(AuditDimensionResult)
                    .filter(AuditDimensionResult.push_log_id.in_(id_chunk))
                    .order_by(AuditDimensionResult.push_log_id.asc(), AuditDimensionResult.id.asc())
                    .all()
                )
            for item in dims:
                dimensions_by_log_id[item.push_log_id].append(item)

        history_by_feedback_id = defaultdict(list)
        if feedback_ids:
            histories = []
            for id_chunk in self._iter_chunks(feedback_ids):
                histories.extend(
                    self.db.query(QCFeedbackHistory)
                    .filter(QCFeedbackHistory.feedback_id.in_(id_chunk))
                    .order_by(QCFeedbackHistory.feedback_id.asc(), QCFeedbackHistory.changed_at.asc())
                    .all()
                )
            for item in histories:
                history_by_feedback_id[item.feedback_id].append(item)

        user_ids = set()
        for _, _, fb, _ in rows:
            if fb and fb.created_by:
                user_ids.add(fb.created_by)
            if fb and fb.assigned_to:
                user_ids.add(fb.assigned_to)
        user_map = {}
        if user_ids:
            users = []
            for id_chunk in self._iter_chunks(list(user_ids)):
                users.extend(self.db.query(User).filter(User.id.in_(id_chunk)).all())
            user_map = {u.id: (u.full_name or u.username or f"#{u.id}") for u in users}

        registry = AuditTypeRegistry()
        result = []
        for log, conclusion, feedback, issue_count in rows:
            audit_type = registry.get_or_default(getattr(log, "audit_type_code", "") or "")
            row_audit_code = getattr(log, "audit_type_code", "") or audit_type.code
            is_lab_exam_row = self._is_lab_exam_export(
                [{"audit_type_code": row_audit_code, "audit_type_name": audit_type.name}],
                row_audit_code,
            )
            dept_ref = None
            if feedback and feedback.dept_id:
                dept_ref = dept_by_id.get(feedback.dept_id)
            elif log.dept:
                dept_ref = dept_by_name.get(str(log.dept or "").strip())
            dept_name = dept_ref.name if dept_ref else (log.dept or "")
            snapshot = extract_patient_snapshot(log)
            raw_sections = extract_raw_record_sections(log)
            if mask_sensitive:
                snapshot = apply_privacy_masking(snapshot, mask_cfg)
            if not snapshot.get("dept_name"):
                snapshot["dept_name"] = dept_name

            focus_items = self._parse_focus_items(getattr(conclusion, "focus_items", "") or "")
            focus_items_text = "\n".join(focus_items) if focus_items else ""
            overall_conclusion = getattr(conclusion, "overall_conclusion", "") or ""
            overall_qc_summary = getattr(conclusion, "overall_qc_summary", "") or ""

            dimension_lines = []
            dimension_structured = {}
            dimension_items = []
            for dim in dimensions_by_log_id.get(log.id, []):
                dim_name = self._normalize_dimension_name(dim.dimension)
                if self._should_exclude_dimension_for_export(dim_name, is_lab_exam_row):
                    continue
                status_text = str(dim.status or "").strip()
                medical_text = str(dim.medical_content or "").strip()
                nursing_text = str(dim.nursing_content or "").strip()
                explanation_text = str(dim.issue_summary or dim.explanation or "").strip()
                dim_value = {
                    "dimension_code": str(dim.dimension_code or "").strip(),
                    "dimension": dim_name,
                    "status": status_text,
                    "severity": self._severity_label(dim.severity or ""),
                    "confidence": dim.confidence if dim.confidence is not None else "",
                    "alert_level": self._alert_level_label(dim.alert_level or ""),
                    "closure_hours": dim.closure_hours or 0,
                    "push_strategy": self._push_strategy_label(dim.push_strategy or ""),
                    "outcome_bucket": self._outcome_bucket_label(dim.outcome_bucket or ""),
                    "medical": medical_text,
                    "nursing": nursing_text,
                    "explanation": explanation_text,
                    "recommendation": str(dim.recommendation or "").strip(),
                }

                if dim_name in dimension_structured:
                    # 同一维度出现多条时合并
                    prev = dimension_structured[dim_name]
                    status_text = " | ".join(filter(None, [prev.get("status", ""), status_text]))
                    medical_text = "\n---\n".join(filter(None, [prev.get("medical", ""), medical_text]))
                    nursing_text = "\n---\n".join(filter(None, [prev.get("nursing", ""), nursing_text]))
                    explanation_text = "\n---\n".join(filter(None, [prev.get("explanation", ""), explanation_text]))

                dimension_structured[dim_name] = {
                    "status": status_text,
                    "medical": medical_text,
                    "nursing": nursing_text,
                    "explanation": explanation_text,
                }
                dimension_items.append(dim_value)

                dimension_lines.append(
                    f"[{dim.dimension}] 状态:{dim.status or ''} 严重度:{dim.severity or ''} | "
                    f"病程:{(dim.medical_content or '').strip()} | 护理:{(dim.nursing_content or '').strip()} | "
                    f"说明:{(dim.issue_summary or dim.explanation or '').strip()} | 建议:{(dim.recommendation or '').strip()}"
                )
            dimensions_text = "\n".join(dimension_lines)

            history_lines = []
            if feedback:
                for item in history_by_feedback_id.get(feedback.id, []):
                    changed_at = item.changed_at.strftime("%Y-%m-%d %H:%M:%S") if item.changed_at else ""
                    history_lines.append(
                        f"{changed_at} | {item.old_status or ''} -> {item.new_status or ''} | {item.change_reason or ''}"
                    )
            history_text = "\n".join(history_lines)

            feedback_text = feedback.feedback_text if feedback else ""
            feedback_status = feedback.status if feedback else "pending"
            assignee = user_map.get(feedback.assigned_to, "") if feedback and feedback.assigned_to else ""
            creator = user_map.get(feedback.created_by, "") if feedback and feedback.created_by else ""

            severity_value = (getattr(conclusion, "severity", "") or log.severity or "")
            alert_level_value = getattr(conclusion, "alert_level", "") or getattr(log, "alert_level", "") or ""
            closure_hours_value = getattr(conclusion, "closure_hours", 0) or 0
            push_strategy_value = getattr(conclusion, "push_strategy", "") or ""
            outcome_bucket_value = getattr(conclusion, "outcome_bucket", "") or ""

            result.append(
                {
                    "log_id": log.id,
                    "audit_type_code": getattr(log, "audit_type_code", "") or audit_type.code,
                    "audit_type_name": audit_type.name,
                    "patient_name": snapshot.get("patient_name", "") or log.patient_name or "",
                    "patient_id": snapshot.get("patient_id", "") or log.patient_id or "",
                    "visit_number": getattr(log, "visit_number", "") or "",
                    "admission_no": snapshot.get("admission_no", "") or getattr(log, "admission_no", "") or "",
                    "dept_name": snapshot.get("dept_name", "") or dept_name,
                    "admission_date": snapshot.get("admission_date", ""),
                    "discharge_date": snapshot.get("discharge_date", ""),
                    "admission_diagnosis": snapshot.get("admission_diagnosis", ""),
                    "is_discharged": snapshot.get("is_discharged", ""),
                    "admission_dept_name": snapshot.get("admission_dept_name", ""),
                    "discharge_dept_name": snapshot.get("discharge_dept_name", ""),
                    "discharge_main_diagnosis": snapshot.get("discharge_main_diagnosis", ""),
                    "surgery": snapshot.get("surgery", ""),
                    "id_card": snapshot.get("id_card", ""),
                    "address": snapshot.get("address", ""),
                    "phone": snapshot.get("phone", ""),
                    "feedback_status": self._status_label(feedback_status),
                    "severity": self._severity_label(severity_value),
                    "alert_level": self._alert_level_label(alert_level_value),
                    "closure_hours": closure_hours_value,
                    "push_strategy": self._push_strategy_label(push_strategy_value),
                    "outcome_bucket": self._outcome_bucket_label(outcome_bucket_value),
                    "query_date": log.query_date or "",
                    "push_time": log.push_time.strftime("%Y-%m-%d %H:%M:%S") if log.push_time else "",
                    "issue_count": issue_count or 0,
                    "overall_conclusion": overall_conclusion,
                    "overall_qc_summary": overall_qc_summary,
                    "focus_items": focus_items_text,
                    "mr_text": log.mr_text or "",
                    "medical_documents_text": raw_sections.get("medical_documents_text", "") or "",
                    "nursing_records_text": raw_sections.get("nursing_records_text", "") or "",
                    "dimensions_text": dimensions_text,
                    "dimensions": dimension_structured,
                    "dimension_items": dimension_items,
                    "dimension_judgments_text": self._format_all_dimension_judgments(dimension_items),
                    "feedback_text": feedback_text or "",
                    "assigned_to": assignee,
                    "created_by": creator,
                    "history_text": history_text,
                }
            )

        return result

    def export_to_csv(
        self,
        role_name: str,
        current_user_dept_id: Optional[int],
        status: Optional[str] = None,
        severity: Optional[str] = None,
        audit_type_code: Optional[str] = None,
        dept_id: Optional[int] = None,
        days: int = 30,
        keyword: Optional[str] = None,
    ) -> bytes:
        """导出 CSV（病例详情）"""
        rows = self._build_case_rows(
            role_name=role_name,
            current_user_dept_id=current_user_dept_id,
            status=status,
            severity=severity,
            audit_type_code=audit_type_code,
            dept_id=dept_id,
            days=days,
            keyword=keyword,
            mask_sensitive=False,
        )
        self.last_export_count = len(rows)

        output = io.StringIO()
        dimension_columns = self._dimension_columns_for_export(rows, audit_type_code)
        canonical_dimension_fieldnames = [name for name, _ in dimension_columns]
        fieldnames = [
            "日志ID",
            "患者姓名",
            "患者ID",
            "身份证号",
            "住址",
            "联系电话",
            "住院号",
            "审计类型编码",
            "审计类型名称",
            "所在科室名称",
            "入院日期",
            "出院日期",
            "入院诊断",
            "是否出院",
            "入院科室名称",
            "出院科室名称",
            "出院主诊断",
            "手术",
            "状态",
            "严重度",
            "预警灯号",
            "闭环时限小时",
            "推送策略",
            "问题分层",
            "查询日期",
            "推送时间",
            "问题数",
            "总体结论",
            "整体质控描述",
            "重点关注项",
            "全部维度判断",
            "原始推送病历文书",
            "原始推送护理记录",
            "当前反馈记录",
            "分配给",
            "创建人",
            "状态变更历史",
        ]
        fieldnames.extend(canonical_dimension_fieldnames)

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for item in rows:
            row_data = {
                "日志ID": item["log_id"],
                "患者姓名": item["patient_name"],
                "患者ID": item["patient_id"],
                "身份证号": item["id_card"],
                "住址": item["address"],
                "联系电话": item["phone"],
                "住院号": item["admission_no"],
                "审计类型编码": item.get("audit_type_code", ""),
                "审计类型名称": item.get("audit_type_name", ""),
                "所在科室名称": item["dept_name"],
                "入院日期": item["admission_date"],
                "出院日期": item["discharge_date"],
                "入院诊断": item["admission_diagnosis"],
                "是否出院": item["is_discharged"],
                "入院科室名称": item["admission_dept_name"],
                "出院科室名称": item["discharge_dept_name"],
                "出院主诊断": item["discharge_main_diagnosis"],
                "手术": item["surgery"],
                "状态": item["feedback_status"],
                "严重度": item["severity"],
                "预警灯号": item.get("alert_level", ""),
                "闭环时限小时": item.get("closure_hours", 0),
                "推送策略": item.get("push_strategy", ""),
                "问题分层": item.get("outcome_bucket", ""),
                "查询日期": item["query_date"],
                "推送时间": item["push_time"],
                "问题数": item["issue_count"],
                "总体结论": item["overall_conclusion"],
                "整体质控描述": item["overall_qc_summary"],
                "重点关注项": item["focus_items"],
                "全部维度判断": item.get("dimension_judgments_text", ""),
                "原始推送病历文书": item.get("medical_documents_text", "") or item.get("mr_text", ""),
                "原始推送护理记录": item.get("nursing_records_text", ""),
                "当前反馈记录": item["feedback_text"],
                "分配给": item["assigned_to"],
                "创建人": item["created_by"],
                "状态变更历史": item["history_text"],
            }
            item_dims = item.get("dimensions") or {}
            for display_name, prefixes in dimension_columns:
                dim_value = self._pick_canonical_dimension_value(item_dims, prefixes)
                row_data[display_name] = self._format_canonical_dimension_cell(dim_value)
            writer.writerow(row_data)

        return output.getvalue().encode("utf-8-sig")

    def export_to_excel(
        self,
        role_name: str,
        current_user_dept_id: Optional[int],
        status: Optional[str] = None,
        severity: Optional[str] = None,
        audit_type_code: Optional[str] = None,
        dept_id: Optional[int] = None,
        days: int = 30,
        keyword: Optional[str] = None,
    ) -> _ExportBytes:
        """导出 Excel；依赖缺失时回退 CSV。返回 (bytes, format) 元组，format 为 'xlsx' 或 'csv'。"""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except Exception:
            # openpyxl 不可用，回退 CSV 导出
            csv_bytes = self.export_to_csv(
                role_name=role_name,
                current_user_dept_id=current_user_dept_id,
                status=status,
                severity=severity,
                audit_type_code=audit_type_code,
                dept_id=dept_id,
                days=days,
                keyword=keyword,
            )
            return _ExportBytes(csv_bytes, "csv")

        rows = self._build_case_rows(
            role_name=role_name,
            current_user_dept_id=current_user_dept_id,
            status=status,
            severity=severity,
            audit_type_code=audit_type_code,
            dept_id=dept_id,
            days=days,
            keyword=keyword,
            mask_sensitive=True,
        )
        self.last_export_count = len(rows)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "病例详情导出"

        columns = [
            ("日志ID", 10),
            ("患者姓名", 12),
            ("患者ID", 14),
            ("身份证号", 20),
            ("地址", 28),
            ("联系电话", 16),
            ("次数", 8),
            ("住院号", 14),
            ("审计类型编码", 18),
            ("审计类型", 20),
            ("入院科室", 14),
            ("出院科室", 14),
            ("入院日期", 12),
            ("出院日期", 12),
            ("推送时间", 19),
            ("严重度", 10),
            ("预警灯号", 10),
            ("闭环时限小时", 12),
            ("推送策略", 12),
            ("问题分层", 12),
            ("是否出院", 10),
            ("出院主诊断", 24),
            ("手术", 24),
            ("总体结论", 36),
            ("整体质控描述", 36),
            ("重点关注项", 30),
            ("全部维度判断", 50),
            ("原始推送病历文书", 50),
            ("原始推送护理记录", 50),
        ]
        dimension_columns = self._dimension_columns_for_export(rows, audit_type_code)
        columns.extend([(name, 42) for name, _ in dimension_columns])

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, (name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        for row_idx, item in enumerate(rows, 2):
            values = [
                item["log_id"],
                item["patient_name"],
                item["patient_id"],
                item.get("id_card", ""),
                item.get("address", ""),
                item.get("phone", ""),
                item.get("visit_number", ""),
                item["admission_no"],
                item.get("audit_type_code", ""),
                item.get("audit_type_name", ""),
                item["admission_dept_name"],
                item["discharge_dept_name"],
                item["admission_date"],
                item["discharge_date"],
                item["push_time"],
                item["severity"],
                item.get("alert_level", ""),
                item.get("closure_hours", 0),
                item.get("push_strategy", ""),
                item.get("outcome_bucket", ""),
                item["is_discharged"],
                item["discharge_main_diagnosis"],
                item["surgery"],
                item["overall_conclusion"],
                item["overall_qc_summary"],
                item["focus_items"],
                item.get("dimension_judgments_text", ""),
                item.get("medical_documents_text", "") or item.get("mr_text", ""),
                item.get("nursing_records_text", ""),
            ]
            item_dims = item.get("dimensions") or {}
            for _, prefixes in dimension_columns:
                dim_value = self._pick_canonical_dimension_value(item_dims, prefixes)
                values.append(self._format_canonical_dimension_cell(dim_value))
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        detail_ws = wb.create_sheet("分类判断明细")
        detail_columns = [
            ("日志ID", 10),
            ("患者姓名", 12),
            ("患者ID", 14),
            ("住院号", 14),
            ("审计类型编码", 18),
            ("审计类型", 20),
            ("类别/维度编码", 18),
            ("类别/维度名称", 24),
            ("判断结果", 12),
            ("严重度", 10),
            ("置信度", 10),
            ("预警灯号", 10),
            ("闭环时限小时", 12),
            ("推送策略", 12),
            ("问题分层", 12),
            ("证据A/病程/首页", 42),
            ("证据B/护理/首次病程", 42),
            ("问题说明", 42),
            ("整改建议", 42),
        ]
        for col_idx, (name, width) in enumerate(detail_columns, 1):
            cell = detail_ws.cell(row=1, column=col_idx)
            cell.value = name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            detail_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        detail_row_idx = 2
        for item in rows:
            for dim in item.get("dimension_items") or []:
                detail_values = [
                    item["log_id"],
                    item["patient_name"],
                    item["patient_id"],
                    item["admission_no"],
                    item.get("audit_type_code", ""),
                    item.get("audit_type_name", ""),
                    dim.get("dimension_code", ""),
                    dim.get("dimension", ""),
                    dim.get("status", ""),
                    dim.get("severity", ""),
                    dim.get("confidence", ""),
                    dim.get("alert_level", ""),
                    dim.get("closure_hours", 0),
                    dim.get("push_strategy", ""),
                    dim.get("outcome_bucket", ""),
                    dim.get("medical", ""),
                    dim.get("nursing", ""),
                    dim.get("explanation", ""),
                    dim.get("recommendation", ""),
                ]
                for col_idx, value in enumerate(detail_values, 1):
                    cell = detail_ws.cell(row=detail_row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                detail_row_idx += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return _ExportBytes(output.getvalue(), "xlsx")
