"""测试 patient_visit_export_service 的海量库集成和 Oracle 回退。"""
import pytest

pytest.importorskip("cryptography")

from app.services import patient_visit_export_service as svc
from app.services.config_parser import ConfigParser


def test_query_progress_notes_from_emr_empty_keys():
    """空 patient_keys 应返回空字典。"""
    result = svc._query_progress_notes_from_emr({"enabled": True}, [])
    assert result == {}


def test_query_discharge_records_from_emr_empty_keys():
    """空 patient_keys 应返回空字典。"""
    result = svc._query_discharge_records_from_emr({"enabled": True}, [])
    assert result == {}


def test_export_fallback_to_oracle_when_vastbase_failed(monkeypatch):
    """海量库查询异常时应自动回退 Oracle。"""
    call_log = []

    def fake_load_config():
        return {
            "oracle": {"host": "test", "port": 1521, "service_name": "orcl", "username": "u", "password_enc": ""},
            "emr_vastbase": {"enabled": True, "host": "", "password_enc": "",
                             "use_for_export_progress": True, "use_for_export_discharge": True,
                             "fallback_to_oracle": True},
        }

    def fake_parse_oracle(cfg):
        return {"host": "test", "port": 1521, "service_name": "orcl", "username": "u", "password": ""}

    def fake_parse_emr(cfg):
        return cfg.get("emr_vastbase", {})

    def fake_get_conn(cfg):
        class FakeConn:
            def close(self): pass
        return FakeConn()

    def fake_query_patient_list(conn):
        return [{"patient_id": "P1", "visit_number": "1", "admission_no": "A1"}]

    def fake_query_patient_basic(conn, pids, vns):
        return {("P1", "1"): {"患者ID": "P1", "次数": "1"}}

    def fake_query_progress_from_emr(emr_cfg, keys):
        call_log.append("emr_progress")
        raise ConnectionError("海量库不可用")

    def fake_query_progress_oracle(conn, pids):
        call_log.append("oracle_progress")
        return {("P1", "1"): [{"event_time": "2026-01-01", "record_name": "首次病程", "content": "内容", "creator": "医生"}]}

    def fake_query_discharge_from_emr(emr_cfg, keys):
        call_log.append("emr_discharge")
        raise ConnectionError("海量库不可用")

    def fake_query_discharge_oracle(conn, pids):
        call_log.append("oracle_discharge")
        return {("P1", "1"): [{"event_time": "2026-01-02", "record_name": "出院记录", "content": "内容", "dept": "内科", "creator": "医生"}]}

    def fake_empty(conn, pids):
        return {}

    def fake_empty2(db, keys):
        return {}

    import app.config as config_mod
    import app.oracle_client as oracle_mod
    monkeypatch.setattr(config_mod, "load_config", fake_load_config)
    monkeypatch.setattr(ConfigParser, "parse_oracle_config", staticmethod(fake_parse_oracle))
    monkeypatch.setattr(ConfigParser, "parse_emr_vastbase_config", staticmethod(fake_parse_emr))
    monkeypatch.setattr(oracle_mod, "get_oracle_connection", fake_get_conn)
    monkeypatch.setattr(svc, "_query_patient_list", fake_query_patient_list)
    monkeypatch.setattr(svc, "_query_patient_basic", fake_query_patient_basic)
    monkeypatch.setattr(svc, "_query_progress_notes_from_emr", fake_query_progress_from_emr)
    monkeypatch.setattr(svc, "_query_progress_notes", fake_query_progress_oracle)
    monkeypatch.setattr(svc, "_query_discharge_records_from_emr", fake_query_discharge_from_emr)
    monkeypatch.setattr(svc, "_query_discharge_records", fake_query_discharge_oracle)
    monkeypatch.setattr(svc, "_query_nursing_records", fake_empty)
    monkeypatch.setattr(svc, "_query_lab_reports", fake_empty)
    monkeypatch.setattr(svc, "_query_exam_reports", fake_empty)
    monkeypatch.setattr(svc, "_query_frontpage_surgery", fake_empty)
    monkeypatch.setattr(svc, "_query_push_logs", fake_empty2)

    class FakeSession:
        pass

    xlsx_bytes, ext = svc.export_patient_visit_summary(FakeSession())
    assert ext == "xlsx"
    assert len(xlsx_bytes) > 0
    assert "emr_progress" in call_log
    assert "oracle_progress" in call_log
    assert "emr_discharge" in call_log
    assert "oracle_discharge" in call_log


def test_export_keeps_discharge_columns():
    """CATEGORY_DEFS 应包含 discharge 类别和正确字段。"""
    assert "discharge" in svc._CATEGORY_DEFS
    defn = svc._CATEGORY_DEFS["discharge"]
    field_keys = [k for _, k in defn["fields"]]
    assert "event_time" in field_keys
    assert "record_name" in field_keys
    assert "content" in field_keys
    assert "dept" in field_keys
    assert "creator" in field_keys


def test_export_uses_vastbase_when_enabled(monkeypatch):
    """海量库启用且查询成功时应使用海量库数据，不调用 Oracle。"""
    call_log = []

    def fake_load_config():
        return {
            "oracle": {"host": "test", "port": 1521, "service_name": "orcl", "username": "u", "password_enc": ""},
            "emr_vastbase": {"enabled": True, "host": "10.0.0.1", "database": "emr", "password_enc": "",
                             "use_for_export_progress": True, "use_for_export_discharge": True,
                             "fallback_to_oracle": True},
        }

    def fake_parse_oracle(cfg):
        return {"host": "test", "port": 1521, "service_name": "orcl", "username": "u", "password": ""}

    def fake_parse_emr(cfg):
        return cfg.get("emr_vastbase", {})

    def fake_get_conn(cfg):
        class FakeConn:
            def close(self): pass
        return FakeConn()

    def fake_query_patient_list(conn):
        return [{"patient_id": "P1", "visit_number": "1", "admission_no": "A1"}]

    def fake_query_patient_basic(conn, pids, vns):
        return {("P1", "1"): {"患者ID": "P1", "次数": "1"}}

    def fake_progress_from_emr(emr_cfg, keys):
        call_log.append("emr_progress")
        return {("P1", "1"): [{"event_time": "2026-01-01", "record_name": "首次病程", "content": "内容", "creator": "医生"}]}

    def fake_discharge_from_emr(emr_cfg, keys):
        call_log.append("emr_discharge")
        return {("P1", "1"): [{"event_time": "2026-01-02", "record_name": "出院记录", "content": "内容", "dept": "内科", "creator": "医生"}]}

    def fake_oracle_progress(conn, pids):
        call_log.append("oracle_progress")
        return {}

    def fake_oracle_discharge(conn, pids):
        call_log.append("oracle_discharge")
        return {}

    def fake_empty(conn, pids):
        return {}

    def fake_empty2(db, keys):
        return {}

    import app.config as config_mod
    import app.oracle_client as oracle_mod
    monkeypatch.setattr(config_mod, "load_config", fake_load_config)
    monkeypatch.setattr(ConfigParser, "parse_oracle_config", staticmethod(fake_parse_oracle))
    monkeypatch.setattr(ConfigParser, "parse_emr_vastbase_config", staticmethod(fake_parse_emr))
    monkeypatch.setattr(oracle_mod, "get_oracle_connection", fake_get_conn)
    monkeypatch.setattr(svc, "_query_patient_list", fake_query_patient_list)
    monkeypatch.setattr(svc, "_query_patient_basic", fake_query_patient_basic)
    monkeypatch.setattr(svc, "_query_progress_notes_from_emr", fake_progress_from_emr)
    monkeypatch.setattr(svc, "_query_discharge_records_from_emr", fake_discharge_from_emr)
    monkeypatch.setattr(svc, "_query_progress_notes", fake_oracle_progress)
    monkeypatch.setattr(svc, "_query_discharge_records", fake_oracle_discharge)
    monkeypatch.setattr(svc, "_query_nursing_records", fake_empty)
    monkeypatch.setattr(svc, "_query_lab_reports", fake_empty)
    monkeypatch.setattr(svc, "_query_exam_reports", fake_empty)
    monkeypatch.setattr(svc, "_query_frontpage_surgery", fake_empty)
    monkeypatch.setattr(svc, "_query_push_logs", fake_empty2)

    class FakeSession:
        pass

    xlsx_bytes, ext = svc.export_patient_visit_summary(FakeSession())
    assert ext == "xlsx"
    assert len(xlsx_bytes) > 0
    assert "emr_progress" in call_log
    assert "emr_discharge" in call_log
    assert "oracle_progress" not in call_log
    assert "oracle_discharge" not in call_log


def test_excel_cell_value_truncates_long_content():
    """超过 32767 字符的单元格值应被自动截断。"""
    long_text = "A" * 40000
    result = svc._excel_cell_value(long_text)
    assert len(result) == 32767


def test_excel_cell_value_preserves_short_content():
    """短内容不应被截断。"""
    short_text = "Hello"
    result = svc._excel_cell_value(short_text)
    assert result == "Hello"


def test_excel_cell_value_handles_none():
    """None 值应返回空字符串。"""
    assert svc._excel_cell_value(None) == ""


def test_excel_cell_value_datetime_field():
    """datetime 字段应格式化为字符串。"""
    from datetime import datetime
    dt = datetime(2026, 5, 28, 10, 30, 0)
    result = svc._excel_cell_value(dt, is_datetime_field=True)
    assert result == "2026-05-28 10:30:00"


def test_format_category_record_skips_empty_fields():
    """空字段不输出，有值字段正常输出，空记录返回空字符串。"""
    defn = {
        "label": "手术",
        "fields": [
            ("手术名称", "operation_name"),
            ("手术日期", "operation_date"),
            ("麻醉方式", "anesthesia_method"),
        ],
    }
    record = {"operation_name": "阑尾切除术", "operation_date": "2026-05-01", "anesthesia_method": ""}
    result = svc._format_category_record(defn, record)
    assert "手术名称：阑尾切除术" in result
    assert "手术日期：2026-05-01" in result
    assert "麻醉方式" not in result


def test_format_category_record_empty_record():
    """空记录应返回空字符串。"""
    defn = {"label": "手术", "fields": [("手术名称", "operation_name")]}
    assert svc._format_category_record(defn, {}) == ""


def test_export_summary_compacts_dynamic_record_columns():
    """动态列应为每条记录一列，字段合并到单元格内。"""
    import io
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl 未安装")

    patient_data = [
        {
            "患者ID": "P1",
            "住院号": "A1",
            "住院次数": "1",
            "患者姓名": "张三",
            "性别": "男",
            "年龄": "50",
            "入院日期": "2026-05-01",
            "出院日期": "2026-05-10",
            "入院科室": "内科",
            "出院科室": "内科",
            "入院诊断": "阑尾炎",
            "出院主诊断": "阑尾炎",
            "surgery": [{"operation_name": "阑尾切除术", "operation_date": "2026-05-02", "anesthesia_method": "全麻", "operation_level": "三级", "wound_healing_grade": "甲"}],
            "progress": [{"event_time": "2026-05-01 10:00:00", "record_name": "首次病程", "content": "患者入院", "creator": "张医生"}],
            "nursing": [{"event_time": "2026-05-01 11:00:00", "record_name": "入院护理", "content": "体温正常", "recorder": "李护士"}],
            "lab": [{"result_time": "2026-05-02 08:00:00", "test_name": "血常规", "result": "正常"}],
            "exam": [{"report_time": "2026-05-02 09:00:00", "exam_class": "B超", "description": "腹部", "impression": "未见异常"}],
            "discharge": [{"event_time": "2026-05-10 10:00:00", "record_name": "出院记录", "content": "治愈出院", "dept": "内科", "creator": "张医生"}],
        }
    ]

    xlsx_bytes = svc._build_excel_with_pushlog(patient_data, max_push=0)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    assert "床号" not in headers
    assert "管床医师" not in headers
    assert "手术1" in headers
    assert "病程1" in headers
    assert "护理1" in headers
    assert "检验1" in headers
    assert "检查1" in headers
    assert "出院记录1" in headers
    assert "手术1_手术名称" not in headers
    assert "病程1_记录时间" not in headers
    assert "护理1_记录时间" not in headers
    assert "检验1_报告时间" not in headers
    assert "检查1_报告时间" not in headers
    assert "出院记录1_记录时间" not in headers

    surgery_col = headers.index("手术1") + 1
    surgery_val = ws.cell(row=2, column=surgery_col).value
    assert "手术名称：阑尾切除术" in surgery_val
    assert "手术日期：" in surgery_val
    assert "麻醉方式：全麻" in surgery_val

    progress_col = headers.index("病程1") + 1
    progress_val = ws.cell(row=2, column=progress_col).value
    assert "记录时间：" in progress_val
    assert "标题：首次病程" in progress_val
    assert "内容：患者入院" in progress_val
    assert "创建人：张医生" in progress_val

    exam_col = headers.index("检查1") + 1
    exam_val = ws.cell(row=2, column=exam_col).value
    assert "报告时间：" in exam_val
    assert "印象：未见异常" in exam_val

    discharge_col = headers.index("出院记录1") + 1
    discharge_val = ws.cell(row=2, column=discharge_col).value
    assert "创建人：张医生" in discharge_val
